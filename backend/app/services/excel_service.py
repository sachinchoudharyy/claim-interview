import uuid
from openpyxl import load_workbook
from app.db.supabase_client import supabase


def generate_case_id():
    return "AUTO-" + str(uuid.uuid4())[:8].upper()


# ✅ COLUMN MAPPING (CORRECT)
COLUMN_MAPPING = {
    "bagic number": "bagic_number",
    "bagic - accused/victim": "accused_victim",
    "investigator name": "investigator_name",
    "accused vehicle no": "accused_vehicle_number",
    "vehicle2(victim1)": "victim_vehicle_number",
    "district": "district",
    "police station": "police_station",
    "fir no": "fir_no",
    "sdrd allocation by investigator date": "allocation_date",
    "fir date": "fir_date",
    "accident date as per fir": "accident_date",
    "bns section": "bns_section",
    "sec tagging": "sec_tagging",
    "case status": "case_status",
    "fraud/ lm/no success": "fraud_flag",
    "ground of fraud lm": "fraud_reason",
    "evidence": "fraud_evidence",
    "remark": "remarks"
}


# ✅ SAFE DATETIME CONVERTER
def convert_value(value):
    if value is None:
        return None

    # convert datetime → string
    if hasattr(value, "isoformat"):
        return value.isoformat()

    # remove formulas like "=DATEDIF(...)"
    if isinstance(value, str) and value.startswith("="):
        return None

    return value


def has_value(val):
    if val is None:
        return False

    val = str(val).strip().lower()
    return val not in ["", "na", "null", "none"]


def process_excel(file_path):

    wb = load_workbook(file_path)
    sheet = wb.active

    headers = [str(cell.value).strip().lower() for cell in sheet[1]]
    print("🔍 NORMALIZED HEADERS:", headers)

    inserted_count = 0

    for row in sheet.iter_rows(min_row=2, values_only=True):

        raw_data = {}
        for header, value in zip(headers, row):
            if header:
                raw_data[header] = convert_value(value)

        print("📦 ROW DATA:", raw_data)

        data = {}

        # map fields
        for excel_col, db_col in COLUMN_MAPPING.items():
            data[db_col] = raw_data.get(excel_col)

        # ✅ CLEAN BAGIC
        bagic = str(data.get("bagic_number")).strip() if data.get("bagic_number") else None


        # 🔥 CLAIM TYPE AUTO-DETECTION (PRODUCTION SAFE)

        accused = data.get("accused_vehicle_number")
        victim = data.get("victim_vehicle_number")

        if has_value(accused) or has_value(victim):
            claim_type = "motor"
        else:
            claim_type = "health"

        print("📌 CLAIM TYPE DETECTED:", claim_type)


        if not bagic or bagic.lower() == "none":
            print("⚠️ Skipping row (invalid bagic_number)")
            continue

        data["bagic_number"] = bagic

        try:
            # ✅ DUPLICATE CHECK
            existing = supabase.table("cases") \
                .select("id") \
                .eq("bagic_number", data["bagic_number"]) \
                .execute()

            if existing.data:
                print(f"⚠️ Duplicate skipped: {data['bagic_number']}")
                continue

            # ✅ INSERT (SAFE VALUES)
            supabase.table("cases").insert({
                "case_id": generate_case_id(),
                "claim_type": claim_type,
                "bagic_number": data.get("bagic_number"),
                "accused_victim": data.get("accused_victim"),
                "investigator_name": data.get("investigator_name"),
                "accused_vehicle_number": data.get("accused_vehicle_number"),
                "victim_vehicle_number": data.get("victim_vehicle_number"),
                "district": data.get("district"),
                "police_station": data.get("police_station"),
                "fir_no": data.get("fir_no"),
                "allocation_date": data.get("allocation_date"),
                "fir_date": data.get("fir_date"),
                "accident_date": data.get("accident_date"),
                "bns_section": data.get("bns_section"),
                "sec_tagging": data.get("sec_tagging"),
                "case_status": (data.get("case_status") or "pending").strip(),
                "fraud_flag": data.get("fraud_flag"),
                "fraud_reason": data.get("fraud_reason"),
                "fraud_evidence": data.get("fraud_evidence"),
                "remarks": data.get("remarks")
            }).execute()

            inserted_count += 1

        except Exception as e:
            print("❌ FAILED ROW:", raw_data)
            print("ERROR:", e)

    print(f"✅ TOTAL ROWS INSERTED: {inserted_count}")


    